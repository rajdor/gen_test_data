import sys
import openpyxl

fname=sys.argv[1]
outputDir=sys.argv[2]


# Open the excel workbook
# for each worksheet
#   save in a file that has the worksheet as the filename
#      skip the header row
#      get data from rows and columns and save to the file

wb = openpyxl.load_workbook(fname)
for sheet in wb.worksheets:
    ws = wb[sheet.title]
    print (ws.title)
    
    outfilename = ws.title.lower() + ".txt"
    
    fout = open(outputDir + outfilename, "w")

    # get max row count
    max_row=ws.max_row
    # get max column count
    max_column=ws.max_column
    
    # iterate over all rows
    for i in range(1,max_row+1):
        if i == 1:
            continue
        
        record = ""
        # iterate over all columns
        colNo = 0
        for j in range(1,max_column+1):      
            colNo = colNo + 1
            cell_obj=ws.cell(row=i,column=j) 
            if colNo > 1:
                record = record + "|"
            record = record + str(cell_obj.value)     
    
        fout.write(record + "\n")
    
    fout.close()